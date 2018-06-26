from openpyxl import Workbook
from openpyxl import load_workbook

def readExcel(filename=None):
    if filename:
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    return wb


def writeExcel(wb):
    # 激活 worksheet
    ws = wb.active
    # 数据可以直接分配到单元格中
    ws['A1'] = 42
    # 可以附加行，从第一列开始附加
    ws.append([1, 2, 3])
    # Python 类型会被自动转换
    import datetime
    ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
    ws1 = wb.create_sheet("Mysheet")  # 插入到最后(default)
    ws2 = wb.create_sheet("Mysheet", 0)  # 插入到最开始的位置
    # 保存文件
    wb.save("/Users/chenchang/Documents/线上修复数据记录/script/test.xlsx")

    return


def writeExcel():
    from openpyxl import Workbook
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active
    # 数据可以直接分配到单元格中
    ws['A1'] = 42
    # 可以附加行，从第一列开始附加
    ws.append([1, 2, 3])
    # Python 类型会被自动转换
    import datetime
    ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
    ws1 = wb.create_sheet("Mysheet")  # 插入到最后(default)
    ws2 = wb.create_sheet("Mysheet", 0)  # 插入到最开始的位置
    # 保存文件
    wb.save("/Users/chenchang/Documents/线上修复数据记录/script/test.xlsx")

    return


def loadconfig(configname):
    with open(configname, 'r') as load_f:
        load_dict = json.load(load_f)
        print("加载配置：")
        print(load_dict)
    return load_dict


def writeconfig(load_dict,configname):
    # load_dict['smallberg'] = [8200, {1: [['Python', 81], ['shirt', 300]]}]
    print("更新配置：")
    print(load_dict)
    with open(configname, "w") as dump_f:
        json.dump(load_dict, dump_f)
    return


# 创建文件并写入
def createFile(file_name,param):
    # 打开文件
    fo = open(file_name, "a")
    print("文件名: ", fo.name)

    # 在文件末尾写入一行
    fo.seek(0, 2)

    if isinstance(param, dict):
        for item in param:
            line = fo.write(str(param[item]) + "\r")
    elif isinstance(param, list):
        line = fo.write(str(param) + "\r")
    elif isinstance(param, tuple):
        line = fo.write(str(param) + "\r")
    elif isinstance(param, str):
        line = fo.write(param + "\r")
    else:
        pass

    # 关闭文件
    fo.close()